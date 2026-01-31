"""Tests for excel_tool - Read and manipulate Excel files."""

import importlib.util
from pathlib import Path
from unittest.mock import patch

import pytest
from fastmcp import FastMCP

from aden_tools.tools.excel_tool.excel_tool import register_tools

openpyxl_available = importlib.util.find_spec("openpyxl") is not None

# Test IDs for sandbox
TEST_WORKSPACE_ID = "test-workspace"
TEST_AGENT_ID = "test-agent"
TEST_SESSION_ID = "test-session"


@pytest.fixture
def excel_tools(mcp: FastMCP, tmp_path: Path):
    """Register all Excel tools and return them as a dict."""
    with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
        register_tools(mcp)
        yield {
            "excel_read": mcp._tool_manager._tools["excel_read"].fn,
            "excel_write": mcp._tool_manager._tools["excel_write"].fn,
            "excel_info": mcp._tool_manager._tools["excel_info"].fn,
        }


@pytest.fixture
def excel_tool_fn(excel_tools):
    """Return excel_read function for backward compatibility."""
    return excel_tools["excel_read"]


@pytest.fixture
def session_dir(tmp_path: Path) -> Path:
    """Create and return the session directory within the sandbox."""
    session_path = tmp_path / TEST_WORKSPACE_ID / TEST_AGENT_ID / TEST_SESSION_ID
    session_path.mkdir(parents=True, exist_ok=True)
    return session_path


@pytest.fixture
def basic_excel(session_dir: Path) -> Path:
    """Create a basic Excel file for testing."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    excel_file = session_dir / "basic.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write headers
    ws["A1"] = "name"
    ws["B1"] = "age"
    ws["C1"] = "city"

    # Write data
    ws["A2"] = "Alice"
    ws["B2"] = 30
    ws["C2"] = "NYC"

    ws["A3"] = "Bob"
    ws["B3"] = 25
    ws["C3"] = "LA"

    ws["A4"] = "Charlie"
    ws["B4"] = 35
    ws["C4"] = "Chicago"

    wb.save(excel_file)
    wb.close()
    return excel_file


@pytest.fixture
def large_excel(session_dir: Path) -> Path:
    """Create a larger Excel file for pagination testing."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    excel_file = session_dir / "large.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "LargeData"

    # Write headers
    ws["A1"] = "id"
    ws["B1"] = "value"

    # Write 100 rows
    for i in range(100):
        ws.cell(row=i + 2, column=1, value=i)
        ws.cell(row=i + 2, column=2, value=i * 10)

    wb.save(excel_file)
    wb.close()
    return excel_file


@pytest.fixture
def multi_sheet_excel(session_dir: Path) -> Path:
    """Create an Excel file with multiple sheets."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    excel_file = session_dir / "multi_sheet.xlsx"
    wb = Workbook()

    # First sheet
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "col1"
    ws1["A2"] = "data1"

    # Second sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "col2"
    ws2["A2"] = "data2"

    # Third sheet
    ws3 = wb.create_sheet("Sales")
    ws3["A1"] = "product"
    ws3["B1"] = "revenue"
    ws3["A2"] = "Widget"
    ws3["B2"] = 1000

    wb.save(excel_file)
    wb.close()
    return excel_file


@pytest.fixture
def empty_excel(session_dir: Path) -> Path:
    """Create an empty Excel file (no data, just headers)."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    excel_file = session_dir / "empty.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"

    # Only headers, no data
    ws["A1"] = "name"
    ws["B1"] = "value"

    wb.save(excel_file)
    wb.close()
    return excel_file


@pytest.mark.skipif(not openpyxl_available, reason="openpyxl not installed")
class TestExcelRead:
    """Tests for excel_read function."""

    def test_read_basic_excel(self, excel_tool_fn, basic_excel, tmp_path):
        """Read a basic Excel file successfully."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["sheet_name"] == "Data"
        assert result["columns"] == ["name", "age", "city"]
        assert result["column_count"] == 3
        assert result["row_count"] == 3
        assert len(result["rows"]) == 3
        assert result["rows"][0] == {"name": "Alice", "age": "30", "city": "NYC"}

    def test_read_with_limit(self, excel_tool_fn, basic_excel, tmp_path):
        """Read Excel with row limit."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                limit=2,
            )

        assert result["success"] is True
        assert result["row_count"] == 2
        assert result["limit"] == 2
        assert len(result["rows"]) == 2
        assert result["rows"][0]["name"] == "Alice"
        assert result["rows"][1]["name"] == "Bob"

    def test_read_with_offset(self, excel_tool_fn, basic_excel, tmp_path):
        """Read Excel with row offset."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                offset=1,
            )

        assert result["success"] is True
        assert result["row_count"] == 2
        assert result["offset"] == 1
        assert result["rows"][0]["name"] == "Bob"
        assert result["rows"][1]["name"] == "Charlie"

    def test_read_with_limit_and_offset(self, excel_tool_fn, large_excel, tmp_path):
        """Read Excel with both limit and offset (pagination)."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="large.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                limit=10,
                offset=50,
            )

        assert result["success"] is True
        assert result["row_count"] == 10
        assert result["offset"] == 50
        assert result["limit"] == 10
        # First row should be id=50
        assert result["rows"][0] == {"id": "50", "value": "500"}

    def test_negative_limit(self, excel_tool_fn, basic_excel, tmp_path):
        """Return error for negative limit."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                limit=-1,
            )

        assert "error" in result
        assert "non-negative" in result["error"].lower()

    def test_negative_offset(self, excel_tool_fn, basic_excel, tmp_path):
        """Return error for negative offset."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                offset=-1,
            )

        assert "error" in result
        assert "non-negative" in result["error"].lower()

    def test_file_not_found(self, excel_tool_fn, session_dir, tmp_path):
        """Return error for non-existent file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="nonexistent.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result
        assert "not found" in result["error"].lower()

    def test_non_xlsx_extension(self, excel_tool_fn, session_dir, tmp_path):
        """Return error for non-Excel file extension."""
        # Create a text file
        txt_file = session_dir / "data.txt"
        txt_file.write_text("name,age\nAlice,30\n")

        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="data.txt",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result
        assert ".xlsx" in result["error"].lower()

    def test_read_specific_sheet(self, excel_tool_fn, multi_sheet_excel, tmp_path):
        """Read a specific sheet by name."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                sheet_name="Sales",
            )

        assert result["success"] is True
        assert result["sheet_name"] == "Sales"
        assert result["columns"] == ["product", "revenue"]
        assert result["rows"][0] == {"product": "Widget", "revenue": "1000"}

    def test_read_invalid_sheet_name(self, excel_tool_fn, multi_sheet_excel, tmp_path):
        """Return error for invalid sheet name."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                sheet_name="NonExistent",
            )

        assert "error" in result
        assert "not found" in result["error"].lower()

    def test_missing_workspace_id(self, excel_tool_fn, basic_excel, tmp_path):
        """Return error when workspace_id is missing."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="basic.xlsx",
                workspace_id="",
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result

    def test_missing_agent_id(self, excel_tool_fn, basic_excel, tmp_path):
        """Return error when agent_id is missing."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id="",
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result

    def test_missing_session_id(self, excel_tool_fn, basic_excel, tmp_path):
        """Return error when session_id is missing."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id="",
            )

        assert "error" in result

    def test_read_unicode_content(self, excel_tool_fn, session_dir, tmp_path):
        """Read Excel with Unicode content."""
        from openpyxl import Workbook

        excel_file = session_dir / "unicode.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "名前"
        ws["B1"] = "年齢"
        ws["C1"] = "都市"
        ws["A2"] = "太郎"
        ws["B2"] = 30
        ws["C2"] = "東京"
        ws["A3"] = "Alice"
        ws["B3"] = 25
        ws["C3"] = "北京"
        wb.save(excel_file)
        wb.close()

        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="unicode.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["columns"] == ["名前", "年齢", "都市"]
        assert result["rows"][0]["名前"] == "太郎"
        assert result["rows"][0]["都市"] == "東京"

    def test_path_traversal_blocked(self, excel_tool_fn, session_dir, tmp_path):
        """Prevent path traversal attacks."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="../../../etc/passwd",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result

    def test_offset_beyond_rows(self, excel_tool_fn, basic_excel, tmp_path):
        """Offset beyond available rows returns empty result."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tool_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                offset=100,
            )

        assert result["success"] is True
        assert result["row_count"] == 0
        assert result["rows"] == []


@pytest.mark.skipif(not openpyxl_available, reason="openpyxl not installed")
class TestExcelWrite:
    """Tests for excel_write function."""

    def test_write_new_excel(self, excel_tools, session_dir, tmp_path):
        """Write a new Excel file successfully."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["name", "age", "city"],
                rows=[
                    {"name": "Alice", "age": "30", "city": "NYC"},
                    {"name": "Bob", "age": "25", "city": "LA"},
                ],
            )

        assert result["success"] is True
        assert result["sheet_name"] == "Sheet1"
        assert result["columns"] == ["name", "age", "city"]
        assert result["column_count"] == 3
        assert result["rows_written"] == 2

        # Verify file was created
        excel_file = session_dir / "output.xlsx"
        assert excel_file.exists()

    def test_write_creates_parent_directories(self, excel_tools, session_dir, tmp_path):
        """Write creates parent directories if needed."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="subdir/nested/output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["id"],
                rows=[{"id": "1"}],
            )

        assert result["success"] is True
        assert (session_dir / "subdir" / "nested" / "output.xlsx").exists()

    def test_write_empty_columns_error(self, excel_tools, session_dir, tmp_path):
        """Return error when columns is empty."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=[],
                rows=[],
            )

        assert "error" in result
        assert "empty" in result["error"].lower()

    def test_write_non_xlsx_extension_error(self, excel_tools, session_dir, tmp_path):
        """Return error for non-Excel file extension."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.xls",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["id"],
                rows=[],
            )

        assert "error" in result
        assert ".xlsx" in result["error"].lower()

    def test_write_filters_extra_columns(self, excel_tools, session_dir, tmp_path):
        """Extra columns in rows are filtered out."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["name"],
                rows=[{"name": "Alice", "extra": "ignored"}],
            )

        assert result["success"] is True

        # Verify by reading back
        excel_file = session_dir / "output.xlsx"
        assert excel_file.exists()

    def test_write_empty_rows(self, excel_tools, session_dir, tmp_path):
        """Write Excel with headers but no rows."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["name", "age"],
                rows=[],
            )

        assert result["success"] is True
        assert result["rows_written"] == 0

        excel_file = session_dir / "output.xlsx"
        assert excel_file.exists()

    def test_write_custom_sheet_name(self, excel_tools, session_dir, tmp_path):
        """Write Excel with custom sheet name."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["id"],
                rows=[{"id": "1"}],
                sheet_name="CustomSheet",
            )

        assert result["success"] is True
        assert result["sheet_name"] == "CustomSheet"

    def test_write_unicode_content(self, excel_tools, session_dir, tmp_path):
        """Write Excel with Unicode content."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="unicode.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["名前", "都市"],
                rows=[{"名前": "太郎", "都市": "東京"}],
            )

        assert result["success"] is True

        excel_file = session_dir / "unicode.xlsx"
        assert excel_file.exists()

    def test_write_numeric_and_boolean_values(self, excel_tools, session_dir, tmp_path):
        """Write Excel with numeric and boolean values."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["id", "price", "active"],
                rows=[
                    {"id": 1, "price": 99.99, "active": True},
                    {"id": 2, "price": 149.5, "active": False},
                ],
            )

        assert result["success"] is True
        assert result["rows_written"] == 2


@pytest.mark.skipif(not openpyxl_available, reason="openpyxl not installed")
class TestExcelInfo:
    """Tests for excel_info function."""

    def test_get_info_basic_excel(self, excel_tools, basic_excel, tmp_path):
        """Get info about a basic Excel file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_info"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["sheet_names"] == ["Data"]
        assert result["active_sheet"] == "Data"
        assert result["columns"] == ["name", "age", "city"]
        assert result["column_count"] == 3
        assert result["total_rows"] == 3
        assert "file_size_bytes" in result
        assert result["file_size_bytes"] > 0

    def test_get_info_multi_sheet(self, excel_tools, multi_sheet_excel, tmp_path):
        """Get info about multi-sheet Excel file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_info"](
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert "Sheet1" in result["sheet_names"]
        assert "Sheet2" in result["sheet_names"]
        assert "Sales" in result["sheet_names"]
        assert result["active_sheet"] == "Sheet1"

    def test_get_info_large_excel(self, excel_tools, large_excel, tmp_path):
        """Get info about a large Excel file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_info"](
                path="large.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["total_rows"] == 100
        assert result["columns"] == ["id", "value"]

    def test_get_info_file_not_found(self, excel_tools, session_dir, tmp_path):
        """Return error when file doesn't exist."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_info"](
                path="nonexistent.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result
        assert "not found" in result["error"].lower()

    def test_get_info_empty_excel(self, excel_tools, empty_excel, tmp_path):
        """Get info about Excel with only headers."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_info"](
                path="empty.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["columns"] == ["name", "value"]
        assert result["total_rows"] == 0

    def test_get_info_non_xlsx_extension_error(self, excel_tools, session_dir, tmp_path):
        """Return error for non-Excel file extension."""
        txt_file = session_dir / "data.txt"
        txt_file.write_text("name\nAlice\n")

        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_info"](
                path="data.txt",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result
        assert ".xlsx" in result["error"].lower()


class TestExcelImportError:
    """Tests for handling missing openpyxl dependency."""

    def test_read_without_openpyxl(self, mcp: FastMCP, tmp_path: Path, monkeypatch):
        """Return error when openpyxl is not installed for read."""
        # Mock openpyxl as not available
        monkeypatch.setattr(
            "aden_tools.tools.excel_tool.excel_tool.importlib.util.find_spec",
            lambda x: None if x == "openpyxl" else True,
        )

        # Reload the module to pick up the mocked import
        import importlib
        from aden_tools.tools import excel_tool as excel_module

        importlib.reload(excel_module)

        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            register_tools(mcp)
            excel_read_fn = mcp._tool_manager._tools["excel_read"].fn

            result = excel_read_fn(
                path="test.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result
        assert "openpyxl" in result["error"].lower()

    def test_write_without_openpyxl(self, mcp: FastMCP, tmp_path: Path, monkeypatch):
        """Return error when openpyxl is not installed for write."""
        # Remove openpyxl from import
        monkeypatch.setattr(
            "builtins.__import__",
            lambda name, *args, **kwargs: None if name == "openpyxl" else __builtins__[name],
        )

        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            register_tools(mcp)
            excel_write_fn = mcp._tool_manager._tools["excel_write"].fn

            result = excel_write_fn(
                path="test.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["id"],
                rows=[],
            )

        assert "error" in result
        assert "openpyxl" in result["error"].lower()
