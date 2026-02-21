import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import CellIsRule

from .schemas import ExcelSchema
from .export_utils import build_export_path


MAX_ROWS = 100000
MAX_SHEETS = 10


MAX_ROWS = 100000
MAX_SHEETS = 10


def generate_excel(schema: ExcelSchema) -> str:

    if len(schema.sheets) > MAX_SHEETS:
        raise ValueError("Too many sheets in Excel file.")

    for sheet in schema.sheets:
        if len(sheet.rows) > MAX_ROWS:
            raise ValueError(
                f"Row limit exceeded in sheet '{sheet.name}'."
            )

    file_path = build_export_path(schema.file_name, "xlsx")

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:

        for sheet in schema.sheets:

            if not sheet.rows:
                continue

            df = pd.DataFrame(sheet.rows)
            df.to_excel(writer, sheet_name=sheet.name, index=False)

            worksheet = writer.sheets[sheet.name]


            #--------------------------------------------------
            # Freeze header
            #---------------------------------------------------
            worksheet.freeze_panes = "A2"

            #--------------------------------------------------
            # Bold header
            #--------------------------------------------------
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            #-------------------------------------------------
            # Auto width
            #-------------------------------------------------
            for column_cells in worksheet.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value else 0
                    for cell in column_cells
                )
                worksheet.column_dimensions[
                    column_cells[0].column_letter
                ].width = max_length + 2


            #-------------------------------------------------
            # Apply formulas
            #-------------------------------------------------
            if sheet.formula_columns:
                df_columns = list(df.columns)

                for row_index, row_data in enumerate(sheet.rows, start=2):
                    for col_name in sheet.formula_columns:

                        if col_name not in df_columns:
                            raise ValueError(
                                f"Formula column '{col_name}' not found."
                            )

                        col_index = df_columns.index(col_name) + 1
                        cell = worksheet.cell(row=row_index, column=col_index)

                        formula_value = row_data.get(col_name)

                        if isinstance(formula_value, str) and formula_value.startswith("="):
                            cell.value = formula_value
                        else:
                            raise ValueError(
                                f"Formula in '{col_name}' must start with '='."
                            )
            #--------------------------------------------------
            # Column formatting
            #--------------------------------------------------
            df_columns = list(df.columns)

            for fmt in sheet.column_formats:

                if fmt.column not in df_columns:
                    raise ValueError(
                        f"Format column '{fmt.column}' not found."
                    )

                col_index = df_columns.index(fmt.column) + 1
                col_letter = worksheet.cell(row=1, column=col_index).column_letter

                if fmt.width:
                    worksheet.column_dimensions[col_letter].width = fmt.width

                for row in worksheet.iter_rows(
                    min_row=2,
                    min_col=col_index,
                    max_col=col_index,
                ):
                    cell = row[0]

                    if fmt.number_format:
                        cell.number_format = fmt.number_format

                    if fmt.alignment:
                        cell.alignment = Alignment(
                            horizontal=fmt.alignment
                        )
            #----------------------------------------------------
            # Conditional formatting
            #----------------------------------------------------
            for cond in sheet.conditional_formats:

                if cond.column not in df_columns:
                    raise ValueError(
                        f"Conditional column '{cond.column}' not found."
                    )

                col_index = df_columns.index(cond.column) + 1
                col_letter = worksheet.cell(row=1, column=col_index).column_letter

                range_str = f"{col_letter}2:{col_letter}{len(df)+1}"

                fill = PatternFill(
                    start_color="FF9999",
                    end_color="FF9999",
                    fill_type="solid"
                )

                operator = "greaterThan" if cond.type == "greater_than" else "lessThan"

                rule = CellIsRule(
                    operator=operator,
                    formula=[str(cond.value)],
                    fill=fill
                )

                worksheet.conditional_formatting.add(range_str, rule)


            #-------------------------------------------------------------
            # Auto filter
            #-------------------------------------------------------------
            if sheet.auto_filter:
                worksheet.auto_filter.ref = worksheet.dimensions


            #-------------------------------------------------------------
            # Charts
            #-------------------------------------------------------------
            for chart_config in sheet.charts:

                if chart_config.chart_type == "line":
                    chart = LineChart()
                else:
                    chart = BarChart()

                chart.title = chart_config.title or "Chart"

                if chart_config.width:
                    chart.width = chart_config.width

                if chart_config.height:
                    chart.height = chart_config.height

                df_columns = list(df.columns)

                x_index = df_columns.index(chart_config.x_column) + 1

                for y_col in chart_config.y_columns:
                    y_index = df_columns.index(y_col) + 1

                    data = Reference(
                        worksheet,
                        min_col=y_index,
                        min_row=1,
                        max_row=len(df) + 1,
                    )

                    chart.add_data(data, titles_from_data=True)

                cats = Reference(
                    worksheet,
                    min_col=x_index,
                    min_row=2,
                    max_row=len(df) + 1,
                )

                chart.set_categories(cats)

                worksheet.add_chart(chart, chart_config.position)

    return str(file_path)
