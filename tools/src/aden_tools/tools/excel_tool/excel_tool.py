"""Excel Tool - Read and manipulate Excel files (.xlsx, .xlsm, .xls)."""

import csv
import os

from fastmcp import FastMCP

from ..file_system_toolkits.security import get_secure_path


def register_tools(mcp: FastMCP) -> None:
    """Register Excel tools with the MCP server."""

    @mcp.tool()
    def excel_read(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        sheet_name: str | None = None,
        limit: int | None = None,
        offset: int = 0,
    ) -> dict:
        """
        Read an Excel file and return its contents.

        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            sheet_name: Name of the sheet to read (None = first sheet)
            limit: Maximum number of rows to return (None = all rows)
            offset: Number of rows to skip from the beginning (after header)

        Returns:
            dict with success status, data, and metadata
        """
        if offset < 0 or (limit is not None and limit < 0):
            return {"error": "offset and limit must be non-negative"}

        try:
            import openpyxl
        except ImportError:
            return {
                "error": ("openpyxl not installed. Install with: pip install openpyxl")
            }

        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)

            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}

            if not path.lower().endswith((".xlsx", ".xlsm", ".xls")):
                return {"error": "File must have .xlsx, .xlsm, or .xls extension"}

            # Load workbook
            wb = openpyxl.load_workbook(secure_path, read_only=True, data_only=True)

            # Get the sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return {
                        "error": f"Sheet '{sheet_name}' not found. "
                        f"Available sheets: {wb.sheetnames}"
                    }
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title

            # Get all rows
            rows_iter = ws.iter_rows(values_only=True)

            # First row is header
            header_row = next(rows_iter, None)
            if not header_row:
                wb.close()
                return {"error": "Excel file is empty (no header row)"}

            # Clean header (convert None to empty string)
            columns = [str(col) if col is not None else "" for col in header_row]

            # Read data rows with offset and limit
            rows = []
            row_num = 0
            for row_values in rows_iter:
                # Skip offset rows
                if row_num < offset:
                    row_num += 1
                    continue

                # Check limit
                if limit is not None and len(rows) >= limit:
                    break

                # Convert row to dict
                row_dict = {}
                for i, col_name in enumerate(columns):
                    value = row_values[i] if i < len(row_values) else None
                    # Convert None to empty string for consistency
                    row_dict[col_name] = value if value is not None else ""

                rows.append(row_dict)
                row_num += 1

            # Count total rows (excluding header)
            total_rows = ws.max_row - 1 if ws.max_row else 0

            wb.close()

            return {
                "success": True,
                "path": path,
                "sheet_name": sheet_name,
                "columns": columns,
                "column_count": len(columns),
                "rows": rows,
                "row_count": len(rows),
                "total_rows": total_rows,
                "offset": offset,
                "limit": limit,
            }

        except Exception as e:
            return {"error": f"Failed to read Excel file: {str(e)}"}

    @mcp.tool()
    def excel_write(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        columns: list[str],
        rows: list[dict],
        sheet_name: str = "Sheet1",
    ) -> dict:
        """
        Write data to a new Excel file.

        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            columns: List of column names for the header
            rows: List of dictionaries, each representing a row
            sheet_name: Name of the sheet to create (default: "Sheet1")

        Returns:
            dict with success status and metadata
        """
        try:
            import openpyxl
        except ImportError:
            return {
                "error": ("openpyxl not installed. Install with: pip install openpyxl")
            }

        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)

            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}

            if not columns:
                return {"error": "columns cannot be empty"}

            # Create parent directories if needed
            parent_dir = os.path.dirname(secure_path)
            if parent_dir:
                os.makedirs(parent_dir, exist_ok=True)

            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Write header
            ws.append(columns)

            # Write data rows
            for row in rows:
                # Extract values in column order
                row_values = [row.get(col, "") for col in columns]
                ws.append(row_values)

            # Save workbook
            wb.save(secure_path)
            wb.close()

            return {
                "success": True,
                "path": path,
                "sheet_name": sheet_name,
                "columns": columns,
                "column_count": len(columns),
                "rows_written": len(rows),
            }

        except Exception as e:
            return {"error": f"Failed to write Excel file: {str(e)}"}

    @mcp.tool()
    def excel_append(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        rows: list[dict],
        sheet_name: str | None = None,
    ) -> dict:
        """
        Append rows to an existing Excel file.

        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            rows: List of dictionaries to append, keys should match existing columns
            sheet_name: Name of the sheet to append to (None = first sheet)

        Returns:
            dict with success status and metadata
        """
        try:
            import openpyxl
        except ImportError:
            return {
                "error": ("openpyxl not installed. Install with: pip install openpyxl")
            }

        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)

            if not os.path.exists(secure_path):
                return {
                    "error": f"File not found: {path}. Use excel_write to create a new file."
                }

            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}

            if not rows:
                return {"error": "rows cannot be empty"}

            # Load workbook
            wb = openpyxl.load_workbook(secure_path)

            # Get the sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return {
                        "error": f"Sheet '{sheet_name}' not found. "
                        f"Available sheets: {wb.sheetnames}"
                    }
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title

            # Read existing columns from first row
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            columns = [str(col) if col is not None else "" for col in first_row]

            # Append new rows
            for row in rows:
                row_values = [row.get(col, "") for col in columns]
                ws.append(row_values)

            # Get new total row count (excluding header)
            total_rows = ws.max_row - 1 if ws.max_row else 0

            # Save workbook
            wb.save(secure_path)
            wb.close()

            return {
                "success": True,
                "path": path,
                "sheet_name": sheet_name,
                "rows_appended": len(rows),
                "total_rows": total_rows,
            }

        except Exception as e:
            return {"error": f"Failed to append to Excel file: {str(e)}"}

    @mcp.tool()
    def excel_info(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
    ) -> dict:
        """
        Get metadata about an Excel file without reading all data.

        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier

        Returns:
            dict with file metadata (sheets, columns per sheet, row counts, file size)
        """
        try:
            import openpyxl
        except ImportError:
            return {
                "error": ("openpyxl not installed. Install with: pip install openpyxl")
            }

        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)

            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}

            if not path.lower().endswith((".xlsx", ".xlsm", ".xls")):
                return {"error": "File must have .xlsx, .xlsm, or .xls extension"}

            # Get file size
            file_size = os.path.getsize(secure_path)

            # Load workbook (read-only for efficiency)
            wb = openpyxl.load_workbook(secure_path, read_only=True, data_only=True)

            # Collect info for each sheet
            sheets_info = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Get header row
                first_row = next(
                    ws.iter_rows(min_row=1, max_row=1, values_only=True), None
                )
                if first_row:
                    columns = [str(col) if col is not None else "" for col in first_row]
                else:
                    columns = []

                # Get row count (excluding header)
                total_rows = ws.max_row - 1 if ws.max_row else 0

                sheets_info.append(
                    {
                        "sheet_name": sheet_name,
                        "columns": columns,
                        "column_count": len(columns),
                        "total_rows": total_rows,
                    }
                )

            wb.close()

            return {
                "success": True,
                "path": path,
                "file_size_bytes": file_size,
                "sheet_count": len(sheets_info),
                "sheets": sheets_info,
            }

        except Exception as e:
            return {"error": f"Failed to get Excel file info: {str(e)}"}

    @mcp.tool()
    def excel_sheet_list(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
    ) -> dict:
        """
        List all sheet names in an Excel workbook.

        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier

        Returns:
            dict with list of sheet names
        """
        try:
            import openpyxl
        except ImportError:
            return {
                "error": ("openpyxl not installed. Install with: pip install openpyxl")
            }

        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)

            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}

            if not path.lower().endswith((".xlsx", ".xlsm", ".xls")):
                return {"error": "File must have .xlsx, .xlsm, or .xls extension"}

            # Load workbook (read-only for efficiency)
            wb = openpyxl.load_workbook(secure_path, read_only=True)

            sheet_names = wb.sheetnames

            wb.close()

            return {
                "success": True,
                "path": path,
                "sheets": sheet_names,
                "sheet_count": len(sheet_names),
            }

        except Exception as e:
            return {"error": f"Failed to list Excel sheets: {str(e)}"}

    @mcp.tool()
    def excel_to_csv(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        output_path: str,
        sheet_name: str | None = None,
    ) -> dict:
        """
        Convert an Excel sheet to CSV format.

        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            output_path: Path for the output CSV file (relative to session sandbox)
            sheet_name: Name of the sheet to convert (None = first sheet)

        Returns:
            dict with success status and metadata
        """
        try:
            import openpyxl
        except ImportError:
            return {
                "error": ("openpyxl not installed. Install with: pip install openpyxl")
            }

        try:
            secure_input_path = get_secure_path(
                path, workspace_id, agent_id, session_id
            )
            secure_output_path = get_secure_path(
                output_path, workspace_id, agent_id, session_id
            )

            if not os.path.exists(secure_input_path):
                return {"error": f"Input file not found: {path}"}

            if not path.lower().endswith((".xlsx", ".xlsm", ".xls")):
                return {"error": "Input file must have .xlsx, .xlsm, or .xls extension"}

            if not output_path.lower().endswith(".csv"):
                return {"error": "Output file must have .csv extension"}

            # Load workbook
            wb = openpyxl.load_workbook(
                secure_input_path, read_only=True, data_only=True
            )

            # Get the sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return {
                        "error": f"Sheet '{sheet_name}' not found. "
                        f"Available sheets: {wb.sheetnames}"
                    }
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title

            # Create output directory if needed
            output_dir = os.path.dirname(secure_output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)

            # Write to CSV
            with open(secure_output_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    # Convert None to empty string
                    cleaned_row = [cell if cell is not None else "" for cell in row]
                    writer.writerow(cleaned_row)

            # Count rows written (excluding header)
            rows_written = ws.max_row - 1 if ws.max_row else 0

            wb.close()

            return {
                "success": True,
                "input_path": path,
                "output_path": output_path,
                "sheet_name": sheet_name,
                "rows_written": rows_written,
            }

        except Exception as e:
            return {"error": f"Failed to convert Excel to CSV: {str(e)}"}
